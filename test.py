import pdfplumber
import openpyxl
import os
import time
from pdf2image import convert_from_path
import pytesseract

# 파일 경로 설정
pdf_file_path = "C:/Users/KECC/Desktop/py/test/rup.pdf"
excel_file_path = "C:/Users/KECC/Desktop/py/test/psv.xlsx"
output_file_path = "C:/Users/KECC/Desktop/py/test/test.xlsx"

# 기존 파일 삭제
if os.path.exists(output_file_path):
    while True:
        try:
            os.remove(output_file_path)
            break
        except PermissionError:
            print("파일이 사용 중입니다. 닫고 다시 시도하세요...")
            time.sleep(3)  # 3초 대기 후 다시 시도

# PDF에서 텍스트 추출
def extract_text_from_pdf(pdf_path):
    extracted_text = []
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                text = page.extract_text()
                if text:
                    extracted_text.append(text)
    except:
        print("PDF에서 텍스트를 추출할 수 없습니다. OCR을 사용합니다.")
        images = convert_from_path(pdf_path)
        extracted_text = [pytesseract.image_to_string(img) for img in images]
    
    return "\n".join(extracted_text)

# PDF에서 데이터 추출
pdf_text = extract_text_from_pdf(pdf_file_path)
print("PDF에서 추출한 텍스트:", pdf_text)

# 텍스트 줄 단위로 분할 후 데이터 추출
pdf_lines = pdf_text.split("\n")
pdf_data = {}

for line in pdf_lines:
    line_parts = line.split(":")
    if len(line_parts) > 1:
        key = line_parts[0].strip()
        value = line_parts[1].strip()
        pdf_data[key] = value

print("추출된 데이터:", pdf_data)

# 엑셀 파일 로드
wb = openpyxl.load_workbook(excel_file_path)

# 기존 시트 삭제 후 새로운 시트 생성
sheet_name = "추출 데이터"
if sheet_name in wb.sheetnames:
    del wb[sheet_name]
ws = wb.create_sheet(sheet_name)

# 헤더 추가
ws.append(["항목", "값"])

# 데이터 입력
for key, value in pdf_data.items():
    ws.append([key, value])

# 수정된 파일 저장
wb.save(output_file_path)

print(f"엑셀 파일 저장 완료: {output_file_path}")
